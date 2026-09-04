# -*- coding: utf-8 -*-
"""
Almacen local del sistema.

Principio de diseño: SQLite es la fuente de verdad, el Excel es una EXPORTACION
que se regenera completa cada vez. Nunca se le agregan filas al .xlsx a mano.
Esto evita duplicados, permite corregir un expediente viejo y elimina toda una
clase de errores de corrupcion del archivo.
"""
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extractor import COLUMNAS, _norm
from rutas import carpeta_datos

BASE = carpeta_datos()
DB = BASE / "expedientes.db"


def ahora():
    """Hora local, no UTC.

    CURRENT_TIMESTAMP de SQLite devuelve UTC. Con el huso argentino (-3), un
    expediente cargado el 30 a las 22:00 quedaria archivado en el mes
    siguiente. Como el archivo por mes depende de esto, se guarda local.
    """
    return datetime.now().isoformat(sep=" ", timespec="seconds")


def mes_de(marca):
    """'2026-09-04 00:30:56' -> '2026-09'"""
    return (marca or "")[:7]


def conectar():
    cx = sqlite3.connect(DB)
    cx.execute("""CREATE TABLE IF NOT EXISTS expedientes (
        expediente TEXT PRIMARY KEY,
        datos      TEXT NOT NULL,
        archivos   TEXT NOT NULL,
        creado     TEXT NOT NULL,
        editado    TEXT NOT NULL,
        mes        TEXT NOT NULL,
        estado     TEXT NOT NULL DEFAULT 'ok',
        observacion TEXT NOT NULL DEFAULT '')""")
    _migrar(cx)
    cx.execute("""CREATE TABLE IF NOT EXISTS carreras (
        clave   TEXT NOT NULL,
        carrera TEXT NOT NULL,
        PRIMARY KEY (clave, carrera))""")
    return cx


def _migrar(cx):
    """Agrega la columna mes a bases creadas antes de esta versión."""
    columnas = {c[1] for c in cx.execute("PRAGMA table_info(expedientes)")}
    # los expedientes marcados para corregir no van al Excel hasta resolverse
    if "estado" not in columnas:
        cx.execute("ALTER TABLE expedientes ADD COLUMN estado TEXT NOT NULL DEFAULT 'ok'")
        cx.execute("ALTER TABLE expedientes ADD COLUMN observacion TEXT NOT NULL DEFAULT ''")
        cx.commit()
    if "mes" in columnas:
        return
    cx.execute("ALTER TABLE expedientes ADD COLUMN mes TEXT NOT NULL DEFAULT ''")
    # las marcas viejas quedaron en UTC: se pasan a local antes de sacar el mes
    for expediente, creado in cx.execute(
            "SELECT expediente, creado FROM expedientes WHERE mes = ''").fetchall():
        try:
            # se descarta el huso al final: el resto del sistema guarda la hora
            # local sin sufijo, y dejarlo daría un "+00:00" que miente
            local = (datetime.fromisoformat(creado)
                     .replace(tzinfo=timezone.utc).astimezone()
                     .replace(tzinfo=None))
            marca = local.isoformat(sep=" ", timespec="seconds")
        except (TypeError, ValueError):
            marca = creado or ahora()
        cx.execute("UPDATE expedientes SET creado = ?, mes = ? WHERE expediente = ?",
                   (marca, mes_de(marca), expediente))
    cx.commit()


# ------------------------------------------------------ tabla de carreras ---
def tabla_carreras():
    """{ 'instituto|materia' : [carreras...] } — la clave incluye el instituto
    porque una misma materia puede dictarse en varios institutos."""
    cx = conectar()
    tabla = {}
    for clave, carrera in cx.execute("SELECT clave, carrera FROM carreras"):
        tabla.setdefault(clave, []).append(carrera)
    cx.close()
    return tabla


def aprender_carrera(instituto, materia, carrera):
    if not (instituto and materia and carrera):
        return
    cx = conectar()
    cx.execute("INSERT OR IGNORE INTO carreras VALUES (?,?)",
               (_norm(f"{instituto}|{materia}"), carrera.strip()))
    cx.commit(); cx.close()


# ---------------------------------------------------------- expedientes -----
def existe(expediente):
    cx = conectar()
    r = cx.execute("SELECT 1 FROM expedientes WHERE expediente=?",
                   (expediente,)).fetchone()
    cx.close()
    return r is not None


def guardar(expediente, valores, archivos, estado="ok", observacion=""):
    """Alta o corrección. Al corregir se conserva el mes de importación
    original: el archivo es por cuándo se cargó, no por cuándo se tocó.

    estado 'observado' = el expediente tiene algo mal y necesita que una
    persona lo arregle; no entra al Excel hasta resolverse.
    """
    marca = ahora()
    cx = conectar()
    cx.execute("""INSERT INTO expedientes
                    (expediente, datos, archivos, creado, editado, mes,
                     estado, observacion)
                  VALUES (?,?,?,?,?,?,?,?)
                  ON CONFLICT(expediente) DO UPDATE SET
                    datos=excluded.datos, archivos=excluded.archivos,
                    editado=excluded.editado, estado=excluded.estado,
                    observacion=excluded.observacion""",
               (expediente, json.dumps(valores, ensure_ascii=False),
                json.dumps(archivos, ensure_ascii=False),
                marca, marca, mes_de(marca), estado, observacion))
    cx.commit(); cx.close()


def obtener(expediente):
    cx = conectar()
    fila = cx.execute("SELECT datos, archivos, creado, editado, mes, estado, "
                      "observacion FROM expedientes WHERE expediente = ?",
                      (expediente,)).fetchone()
    cx.close()
    if not fila:
        return None
    datos, archivos, creado, editado, mes, estado, observacion = fila
    return {"expediente": expediente, "valores": json.loads(datos),
            "archivos": json.loads(archivos), "creado": creado,
            "editado": editado, "mes": mes, "estado": estado,
            "observacion": observacion}


def eliminar(expediente):
    cx = conectar()
    cx.execute("DELETE FROM expedientes WHERE expediente = ?", (expediente,))
    borrados = cx.total_changes
    cx.commit(); cx.close()
    return borrados


def contexto_excel(expediente):
    """Donde va a caer esta fila en el Excel, para poder mostrarlo antes de
    guardar. Ver el destino evita la duda de si se duplica o se pisa."""
    buenos = [f for f in _filas() if f[6] == "ok"]
    fila = next((i for i, f in enumerate(buenos, start=1)
                 if f[0] == expediente), None)
    ultimos = [{"expediente": e,
                "apellido": v.get("apellido", ""), "nombre": v.get("nombre", ""),
                "dni": v.get("dni", ""), "carrera": v.get("carrera", ""),
                "fecha_alta": v.get("fecha_alta", ""),
                "fila": i, "es_este": e == expediente}
               for i, (e, v, *_) in enumerate(buenos, start=1)][-2:]
    return {"archivo": (BASE / "expedientes.xlsx").name,
            "hoja": "Todos",
            "total": len(buenos),
            "fila": fila,                 # None si es nuevo
            "ultimos": ultimos}


def historial():
    """Los expedientes agrupados por mes de importación, del más nuevo al más
    viejo."""
    meses = {}
    for (expediente, valores, _archivos, creado, editado, mes,
         estado, observacion) in _filas():
        meses.setdefault(mes, []).append({
            "expediente": expediente,
            "docente": " ".join(x for x in (valores.get("apellido"),
                                            valores.get("nombre")) if x),
            "instituto": valores.get("instituto", ""),
            "materia": valores.get("materia", ""),
            "creado": creado,
            "editado": editado if editado != creado else "",
            "estado": estado,
            "observacion": observacion,
        })
    return [{"mes": m, "expedientes": sorted(e, key=lambda x: x["creado"], reverse=True)}
            for m, e in sorted(meses.items(), reverse=True)]


def _filas():
    cx = conectar()
    filas = [(e, json.loads(d), json.loads(a), c, ed, m, es, ob)
             for e, d, a, c, ed, m, es, ob in cx.execute(
                 "SELECT expediente, datos, archivos, creado, editado, mes, "
                 "estado, observacion FROM expedientes ORDER BY creado")]
    cx.close()
    return filas


def listar(solo_ok=False):
    return [(e, v, a, c) for e, v, a, c, _ed, _m, es, _ob in _filas()
            if not solo_ok or es == "ok"]


# ------------------------------------------------------- exportar Excel -----
ENCABEZADO = PatternFill("solid", fgColor="1F3864")
REVISAR    = PatternFill("solid", fgColor="FFF2CC")   # celda vacia / sin dato


MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def nombre_hoja(mes):
    """'2026-09' -> 'Septiembre 2026'"""
    try:
        anio, numero = mes.split("-")
        return f"{MESES[int(numero) - 1].capitalize()} {anio}"
    except (ValueError, IndexError):
        return mes or "Sin fecha"


def _escribir_hoja(ws, filas, con_importacion, extras=()):
    etiquetas = [etiqueta for _, etiqueta in COLUMNAS]
    if con_importacion:
        etiquetas.append("Importado")
    etiquetas += list(extras)
    ws.append(etiquetas)
    for i in range(1, len(etiquetas) + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = ENCABEZADO
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for _expediente, valores, _archivos, creado, *resto in filas:
        fila = [valores.get(k, "") for k, _ in COLUMNAS]
        if con_importacion:
            fila.append(creado)
        fila += list(resto)          # motivo, cuando hay hoja de observados
        ws.append(fila)
        r = ws.max_row
        for i, (clave, _) in enumerate(COLUMNAS, start=1):
            if not str(valores.get(clave, "")).strip():
                ws.cell(row=r, column=i).fill = REVISAR   # resaltar lo que falta

    anchos = {"Instituto": 18, "N° de expediente": 34, "Descripción": 40,
              "Materia": 42, "Carrera": 38, "Cantidad de firmas": 18,
              "Importado": 20, "Qué hay que corregir": 60}
    for i, etiqueta in enumerate(etiquetas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(etiqueta, 16)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def exportar_excel(destino=None):
    """Regenera el Excel completo desde SQLite.

    Una hoja "Todos" con el historial entero, y despues una hoja por mes de
    importacion, de la mas reciente a la mas vieja.
    """
    destino = Path(destino or BASE / "expedientes.xlsx")
    filas = _filas()
    # Los marcados para corregir NO entran a las hojas normales: el expediente
    # tiene algo mal y no debe seguir su curso hasta que alguien lo arregle.
    buenos = [f for f in filas if f[6] == "ok"]
    observados = [f for f in filas if f[6] != "ok"]

    wb = Workbook()
    todos = wb.active
    todos.title = "Todos"
    _escribir_hoja(todos, [(e, v, a, c) for e, v, a, c, *_ in buenos], True)

    por_mes = {}
    for e, v, a, c, _ed, mes, _es, _ob in buenos:
        por_mes.setdefault(mes, []).append((e, v, a, c))
    for mes in sorted(por_mes, reverse=True):
        _escribir_hoja(wb.create_sheet(nombre_hoja(mes)), por_mes[mes], False)

    # pero quedan a la vista en su propia hoja: son trabajo pendiente
    if observados:
        _escribir_hoja(wb.create_sheet("Para corregir"),
                       [(e, v, a, c, ob) for e, v, a, c, _ed, _m, _es, ob in observados],
                       True, extras=("Qué hay que corregir",))

    wb.save(destino)
    return destino
